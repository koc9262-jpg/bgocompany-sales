#!/usr/bin/env python3
import openpyxl, os, re, json, sys
from datetime import date

script_dir = os.path.dirname(os.path.abspath(__file__))
data_dir = os.path.join(script_dir, 'data')

if not os.path.exists(data_dir):
    print("ERROR: data/ not found", file=sys.stderr); sys.exit(1)

all_xlsx = [f for f in os.listdir(data_dir)
            if re.match(r'\d{4}_b_', f) and f.endswith('.xlsx')]
if not all_xlsx:
    print("INFO: no xlsx in data/ - skip", file=sys.stderr); sys.exit(0)

prefixes = sorted(set(re.match(r'(\d{4})_', f).group(1)
                  for f in all_xlsx if re.match(r'(\d{4})_', f)))
prefix = prefixes[-1]
print(f"Parsing prefix: {prefix}", file=sys.stderr)

files = sorted([f for f in all_xlsx if f.startswith(prefix)])
branch_files = {}
for f in files:
    m = re.match(rf'({prefix}_b_\d+_\S+?)(?:\s*\(\d+\))?\.xlsx', f)
    key = m.group(1) if m else re.sub(r'\.xlsx$', '', f)
    branch_files[key] = f

b5_exists = any(re.match(rf'{prefix}_b_5', k) for k in branch_files)
if not b5_exists:
    branch_files[f'{prefix}_b_5_jangan'] = None

def safe(row, i):
    try: return row[i]
    except IndexError: return None

def pct(v):
    if v is None: return 0.0
    v = float(v)
    return round(v * 100, 1) if abs(v) < 10 else round(v, 1)

results = []
for key in sorted(branch_files.keys()):
    fname = branch_files[key]
    if fname is None:
        branch_name = key.split('_', 3)[-1]
        results.append({'branch': branch_name, 'target_fc': 0, 'target_pt': 0, 'target_tot': 0,
            'actual_fc': 0, 'actual_pt': 0, 'actual_tot': 0,
            'rate_fc': 0.0, 'rate_pt': 0.0, 'rate_tot': 0.0,
            'fc_new_cnt': 0, 'fc_new_amt': 0, 'fc_re_cnt': 0, 'fc_re_amt': 0,
            'pt_new_cnt': 0, 'pt_new_amt': 0, 'pt_re_cnt': 0, 'pt_re_amt': 0,
            'total_card': 0, 'cash_out': 0, 'account': 0, 'deduction': 0, 'real_cash_out': 0,
            'ranks': []})
        continue
    try:
        wb = openpyxl.load_workbook(os.path.join(data_dir, fname), data_only=True)
    except Exception:
        wb = openpyxl.load_workbook(os.path.join(data_dir, fname), data_only=True, read_only=True)
    ws = wb['sheet1'] if 'sheet1' in [s.lower() for s in wb.sheetnames] else wb.active
    for sname in wb.sheetnames:
        if sname == '매출':
            ws = wb[sname]
            break
    rows = [[cell.value for cell in row] for row in ws.iter_rows()]
    r7 = rows[7] if len(rows) > 7 else []
    r10 = rows[10] if len(rows) > 10 else []
    ranks = []
    for sname in wb.sheetnames:
        if '순위' in sname:
            ws2 = wb[sname]
            for row in list(ws2.iter_rows())[3:]:
                rv = [c.value for c in row]
                nm, amt = safe(rv,3), safe(rv,4)
                if isinstance(nm, str) and nm.strip() and isinstance(amt, (int, float)) and amt > 0:
                    ranks.append({'name': nm.strip(), 'amount': int(amt)})
    results.append({'branch': safe(rows[0],20) if rows else None,
        'target_fc': int(safe(rows[2],2) or 0) if len(rows)>2 else 0,
        'target_pt': int(safe(rows[2],4) or 0) if len(rows)>2 else 0,
        'target_tot': int(safe(rows[2],6) or 0) if len(rows)>2 else 0,
        'actual_fc': int(safe(rows[3],2) or 0) if len(rows)>3 else 0,
        'actual_pt': int(safe(rows[3],4) or 0) if len(rows)>3 else 0,
        'actual_tot': int(safe(rows[3],6) or 0) if len(rows)>3 else 0,
        'rate_fc': pct(safe(rows[4],2)) if len(rows)>4 else 0.0,
        'rate_pt': pct(safe(rows[4],4)) if len(rows)>4 else 0.0,
        'rate_tot': pct(safe(rows[4],6)) if len(rows)>4 else 0.0,
        'fc_new_cnt': int(safe(r7,10) or 0), 'fc_new_amt': int(safe(r7,11) or 0),
        'fc_re_cnt': int(safe(r7,12) or 0), 'fc_re_amt': int(safe(r7,13) or 0),
        'pt_new_cnt': int(safe(r7,17) or 0), 'pt_new_amt': int(safe(r7,18) or 0),
        'pt_re_cnt': int(safe(r7,19) or 0), 'pt_re_amt': int(safe(r7,20) or 0),
        'total_card': int(safe(r10,12) or 0), 'cash_out': int(safe(r10,14) or 0),
        'account': int(safe(r10,16) or 0), 'deduction': int(safe(r10,18) or 0),
        'real_cash_out': int(safe(r10,20) or 0), 'ranks': ranks})
    print(f"  OK {fname}", file=sys.stderr)

today_str = date.today().strftime('%Y%m%d')
data_json = json.dumps(results, ensure_ascii=False)

html_path = os.path.join(script_dir, 'index.html')
html = open(html_path).read()
html = re.sub(r'const D_RAW = .*?;', f'const D_RAW = {data_json};', html)
html = re.sub(r'(\d{4})\xb7(\d{2})\xb7(\d{2})', today_str, html)
html = re.sub(r'20\d\d\xb477 \d\d\xb477 \d\d\xc77c', '', html)

year = date.today().year
month = date.today().month
day = date.today().day
html = re.sub(r'20\d\d년 \d+월 \d+일', f'{year}년 {month:02d}월 {day:02d}일', html)

with open(html_path, 'w', encoding='utf-8') as f:
    f.write(html)
print(f'Done: {html_path}', file=sys.stderr)
