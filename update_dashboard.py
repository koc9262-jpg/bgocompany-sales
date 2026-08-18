#!/usr/bin/env python3
"""
í´ë©ì´í¬/ìì¤íì´ ë§¤ì¶ ëìë³´ë ìë ìë°ì´í¸ ì¤í¬ë¦½í¸ (GitHub Actions ì©)
data/ í´ëì xlsx íì¼ì íì±íì¬ index.html ì ìì±í©ëë¤.
"""
import openpyxl, os, re, json, sys
from datetime import date

script_dir = os.path.dirname(os.path.abspath(__file__))
data_dir = os.path.join(script_dir, 'data')

if not os.path.exists(data_dir):
    print("ERROR: data/ í´ë ìì", file=sys.stderr); sys.exit(1)

all_xlsx = [f for f in os.listdir(data_dir)
            if re.match(r'\d{4}_b_', f) and f.endswith('.xlsx')]
if not all_xlsx:
    print("INFO: data/ 폴더에 xlsx 파일 없음 - 스킵", file=sys.stderr); sys.exit(0)

prefixes = sorted(set(re.match(r'(\d{4})_', f).group(1)
                  for f in all_xlsx if re.match(r'(\d{4})_', f)))
prefix = prefixes[-1]
print(f"íì± ì: {prefix}", file=sys.stderr)

files = sorted([f for f in all_xlsx if f.startswith(prefix)])
branch_files = {}
for f in files:
    m = re.match(rf'({prefix}_b_\d+_\S+?)(?:\s*\(\d+\))?\.xlsx', f)
    key = m.group(1) if m else f
    branch_files[key] = f

b5_exists = any(re.match(rf'{prefix}_b_5_', k) for k in branch_files)
if not b5_exists:
    branch_files[f'{prefix}_b_5_ì¥ìì '] = None

results = []
for key in sorted(branch_files.keys()):
    fname = branch_files[key]
    if fname is None:
        branch_name = key.split('_', 3)[-1]
        results.append({'branch': branch_name, 'target_fc': 0, 'target_pt': 0, 'target_tot': 0,
            'actual_fc': 0, 'actual_pt': 0, 'actual_tot': 0,
            'rate_fc': 0.0, 'rate_pt': 0.0, 'rate_tot': 0.0,
            'fc_new_cnt': 0, 'fc_new_amt': 0, 'fc_re_cnt': 0, 'fc_re_amt': 0,
            'pt_new_cnt': 0, 'pt_new_amt': 0, 'pt_re_cnt': 0, 'pt_re_amt': 0,
            'total_card': 0, 'cash_out': 0, 'account': 0, 'deduction': 0, 'real_cash_out': 0,
            'ranks': []})
        continue
    try:
        wb = openpyxl.load_workbook(os.path.join(data_dir, fname), data_only=True)
    except Exception:
        wb = openpyxl.load_workbook(os.path.join(data_dir, fname), data_only=True, read_only=True)
    ws = wb['ë§¤ì¶'] if 'ë§¤ì¶' in wb.sheetnames else wb.active
    rows = [[cell.value for cell in row] for row in ws.iter_rows()]
    def pct(v):
        if v is None: return 0.0
        v = float(v)
        return round(v * 100, 1) if abs(v) < 10 else round(v, 1)
    r7 = rows[7]; r10 = rows[10]
    ranks = []
    if 'ë§¤ì¶ìì' in wb.sheetnames:
        ws2 = wb['ë§¤ì¶ìì']
        for row in list(ws2.iter_rows())[3:]:
            rv = [c.value for c in row]
            nm, amt = rv[3], rv[4]
            if isinstance(nm, str) and nm.strip() and isinstance(amt, (int, float)) and amt > 0:
                ranks.append({'name': nm.strip(), 'amount': int(amt)})
    results.append({'branch': rows[0][20],
        'target_fc': int(rows[2][2] or 0), 'target_pt': int(rows[2][4] or 0), 'target_tot': int(rows[2][6] or 0),
        'actual_fc': int(rows[3][2] or 0), 'actual_pt': int(rows[3][4] or 0), 'actual_tot': int(rows[3][6] or 0),
        'rate_fc': pct(rows[4][2]), 'rate_pt': pct(rows[4][4]), 'rate_tot': pct(rows[4][6]),
        'fc_new_cnt': int(r7[10] or 0), 'fc_new_amt': int(r7[11] or 0),
        'fc_re_cnt': int(r7[12] or 0), 'fc_re_amt': int(r7[13] or 0),
        'pt_new_cnt': int(r7[17] or 0), 'pt_new_amt': int(r7[18] or 0),
        'pt_re_cnt': int(r7[19] or 0), 'pt_re_amt': int(r7[20] or 0),
        'total_card': int(r10[12] or 0), 'cash_out': int(r10[14] or 0),
        'account': int(r10[16] or 0), 'deduction': int(r10[18] or 0),
        'real_cash_out': int(r10[20] or 0), 'ranks': ranks})
    print(f"  â {fname}", file=sys.stderr)

today_str = date.today().strftime('%Yë %mì %dì¼')
data_json = json.dumps(results, ensure_ascii=False)

html = open(os.path.join(script_dir, 'index.html')).read()
import re as _re
html = _re.sub(r'const D_RAW = .*?;', f'const D_RAW = {data_json};', html)
html = _re.sub(r"ê¸°ì¤ì¼:.*?Â·", f'ê¸°ì¤ì¼: {today_str} Â·', html)

output_path = os.path.join(script_dir, 'index.html')
with open(output_path, 'w', encoding='utf-8') as f:
    f.write(html)
print(f'ìë£: {output_path}', file=sys.stderr)
